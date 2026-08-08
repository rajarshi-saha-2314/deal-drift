"""
DealDrift Excel report export.

Builds a client-deliverable-style .xlsx via openpyxl: a Summary sheet (KPIs +
top current deals, with a discount bar chart -- real data, always available),
a Price History sheet (full scraped history), and a Forecasts sheet (N-day
forecast table + line chart per ASIN with enough history).

Data caveat (2026-08-08): as of this writing, no ASIN yet has the
modeling.train_model.MIN_OBSERVATIONS daily observations needed for a real
forecast (the scheduler hasn't been running long enough). The Forecasts
sheet reflects that honestly -- it does not fabricate a chart. Re-run this
export once real forecasts exist; no code changes needed, it picks them up
automatically.
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.db_config import get_connection
from modeling.predict import forecast_price
from modeling.train_model import MIN_OBSERVATIONS, get_daily_price_series

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(size=10, italic=True, color="666666")


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Data pulls (raw SQL, no ORM)
# ---------------------------------------------------------------------------

_LATEST_SNAPSHOT_SQL = """
    SELECT p.asin, p.title, p.category, ph.price, ph.list_price, ph.discount_pct,
           ph.currency, ph.rating, ph.review_count, ph.scraped_at
    FROM price_history ph
    JOIN products p ON p.asin = ph.asin
    WHERE ph.price_history_id IN (
        SELECT MAX(price_history_id) FROM price_history GROUP BY asin
    )
    ORDER BY ph.discount_pct DESC
"""

_FULL_HISTORY_SQL = """
    SELECT p.asin, p.title, ph.scraped_at, ph.price, ph.list_price, ph.discount_pct, ph.is_outlier
    FROM price_history ph
    JOIN products p ON p.asin = ph.asin
    ORDER BY p.asin, ph.scraped_at
"""


def get_latest_snapshot() -> list[dict]:
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(_LATEST_SNAPSHOT_SQL)
            return cursor.fetchall()
    finally:
        conn.close()


def get_full_history() -> list[dict]:
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(_FULL_HISTORY_SQL)
            return cursor.fetchall()
    finally:
        conn.close()


def get_all_asins() -> list[str]:
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT DISTINCT asin FROM price_history;")
            return [row["asin"] for row in cursor.fetchall()]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_summary_sheet(wb: Workbook, snapshot: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "DealDrift -- Price Intelligence Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT

    n_products = len(snapshot)
    avg_discount = sum(r["discount_pct"] or 0 for r in snapshot) / n_products if n_products else 0
    top_discount = max((r["discount_pct"] or 0 for r in snapshot), default=0)

    ws["A4"] = "Products tracked:"
    ws["B4"] = n_products
    ws["A5"] = "Average discount (latest snapshot):"
    ws["B5"] = round(avg_discount, 1)
    ws["A6"] = "Top current discount:"
    ws["B6"] = f"{top_discount:.1f}%"
    for row in (4, 5, 6):
        ws[f"A{row}"].font = Font(bold=True)

    header_row = 8
    headers = ["Title", "Category", "Price", "List Price", "Discount %", "Rating", "Reviews", "Snapshot Time"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    _style_header_row(ws, header_row, len(headers))

    top_n = sorted(snapshot, key=lambda r: r["discount_pct"] or 0, reverse=True)[:15]
    for i, row in enumerate(top_n, start=header_row + 1):
        ws.cell(row=i, column=1, value=row["title"][:80] if row["title"] else "")
        ws.cell(row=i, column=2, value=row["category"])
        ws.cell(row=i, column=3, value=float(row["price"]) if row["price"] is not None else None)
        ws.cell(row=i, column=4, value=float(row["list_price"]) if row["list_price"] is not None else None)
        ws.cell(row=i, column=5, value=float(row["discount_pct"]) if row["discount_pct"] is not None else None)
        ws.cell(row=i, column=6, value=float(row["rating"]) if row["rating"] is not None else None)
        ws.cell(row=i, column=7, value=row["review_count"])
        ws.cell(row=i, column=8, value=row["scraped_at"].strftime("%Y-%m-%d %H:%M") if row["scraped_at"] else "")

    _autofit_columns(ws, {"A": 45, "B": 20, "C": 10, "D": 10, "E": 12, "F": 8, "G": 10, "H": 18})
    ws.freeze_panes = f"A{header_row + 1}"

    # Discount bar chart -- real data, always meaningful even without forecasts.
    if top_n:
        chart = BarChart()
        chart.title = "Top Discounts (Latest Snapshot)"
        chart.y_axis.title = "Discount %"
        chart.x_axis.title = "Product"
        data = Reference(ws, min_col=5, min_row=header_row, max_row=header_row + len(top_n))
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(top_n))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 24, 12
        ws.add_chart(chart, f"J{header_row}")


def build_price_history_sheet(wb: Workbook, history: list[dict]) -> None:
    ws = wb.create_sheet("Price History")
    headers = ["ASIN", "Title", "Scraped At", "Price", "List Price", "Discount %", "Outlier?"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for i, row in enumerate(history, start=2):
        ws.cell(row=i, column=1, value=row["asin"])
        ws.cell(row=i, column=2, value=row["title"][:80] if row["title"] else "")
        ws.cell(row=i, column=3, value=row["scraped_at"].strftime("%Y-%m-%d %H:%M") if row["scraped_at"] else "")
        ws.cell(row=i, column=4, value=float(row["price"]) if row["price"] is not None else None)
        ws.cell(row=i, column=5, value=float(row["list_price"]) if row["list_price"] is not None else None)
        ws.cell(row=i, column=6, value=float(row["discount_pct"]) if row["discount_pct"] is not None else None)
        ws.cell(row=i, column=7, value="Yes" if row["is_outlier"] else "No")

    _autofit_columns(ws, {"A": 14, "B": 45, "C": 18, "D": 10, "E": 10, "F": 12, "G": 10})
    ws.freeze_panes = "A2"


def build_forecasts_sheet(wb: Workbook, asins: list[str]) -> None:
    ws = wb.create_sheet("Forecasts")
    ws["A1"] = "7-Day Price Forecasts"
    ws["A1"].font = TITLE_FONT

    fittable = [a for a in asins if len(get_daily_price_series(a)) >= MIN_OBSERVATIONS]

    if not fittable:
        ws["A3"] = (
            f"No product currently has {MIN_OBSERVATIONS}+ daily price observations, "
            "so there are no real forecasts to show yet."
        )
        ws["A4"] = (
            "This sheet will populate automatically once data_extraction/scheduler.py "
            "has accumulated enough daily history -- re-run this export at that point."
        )
        ws["A3"].font = Font(italic=True, color="B00000")
        ws["A4"].font = Font(italic=True, color="666666")
        _autofit_columns(ws, {"A": 100})
        return

    row_cursor = 3
    for asin in fittable:
        forecast_df = forecast_price(asin, days=7, model_type="sarima")
        if forecast_df is None or forecast_df.empty:
            continue

        ws.cell(row=row_cursor, column=1, value=f"ASIN: {asin}").font = Font(bold=True)
        row_cursor += 1
        header_row = row_cursor
        headers = ["Date", "Forecast", "Lower CI", "Upper CI"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header)
        _style_header_row(ws, header_row, len(headers))
        row_cursor += 1

        start_data_row = row_cursor
        for _, fr in forecast_df.iterrows():
            ws.cell(row=row_cursor, column=1, value=fr["date"].strftime("%Y-%m-%d"))
            ws.cell(row=row_cursor, column=2, value=round(float(fr["forecast"]), 2))
            ws.cell(row=row_cursor, column=3, value=round(float(fr["lower_ci"]), 2))
            ws.cell(row=row_cursor, column=4, value=round(float(fr["upper_ci"]), 2))
            row_cursor += 1
        end_data_row = row_cursor - 1

        chart = LineChart()
        chart.title = f"{asin} -- 7-Day Forecast"
        data = Reference(ws, min_col=2, max_col=4, min_row=header_row, max_row=end_data_row)
        cats = Reference(ws, min_col=1, min_row=start_data_row, max_row=end_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 18, 8
        ws.add_chart(chart, f"F{header_row}")

        row_cursor += 16  # leave room for the chart before the next ASIN's table

    _autofit_columns(ws, {"A": 14, "B": 12, "C": 12, "D": 12})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def export_report(output_path: str = "reports/dealdrift_report.xlsx") -> str:
    snapshot = get_latest_snapshot()
    if not snapshot:
        raise RuntimeError("No price_history data found -- run the scraper before exporting a report.")
    history = get_full_history()
    asins = get_all_asins()

    wb = Workbook()
    build_summary_sheet(wb, snapshot)
    build_price_history_sheet(wb, history)
    build_forecasts_sheet(wb, asins)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Report saved to %s", path.resolve())
    return str(path.resolve())


if __name__ == "__main__":
    output = export_report()
    print(f"Report written to: {output}")
